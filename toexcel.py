import openpyxl
from openpyxl import load_workbook
from datetime import datetime

def populate_excel(account_number, form_data, template_path='data1.xlsx'):
    try:
        wb = load_workbook(template_path)
        sheet = wb.active
        
        # Find the row with the specified account number
        account_number = int(account_number)
        account_number_column = 3
        row_with_account = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=account_number_column, max_col=account_number_column):
            if row[0].value == account_number:
                row_with_account = row[0].row
                break

        if row_with_account is None:
            print(f"Account number {account_number} not found in the Excel sheet.")
            return

        # Mapping form data to column indices
        column_indices = {
            'Loan Application Date': 23,
            'Purpose of the Loan': 24,
            'Address/Location': 25,
            'Business Financed': 26,
            'Group Name': 27,
            'Reason for Default (Summarised)': 28,
            'Detailed Reason for Default': 29
        }

        # Writing form data to Excel for the specified account number
        for field, value in form_data.items():
            if field in column_indices:
                column_index = column_indices[field]
                sheet.cell(row=row_with_account, column=column_index, value=value)

        # Adding Loan Application Date if available
        if 'Loan Application Date' in column_indices:
            loan_date_index = column_indices['Loan Application Date']
            sheet.cell(row=row_with_account, column=loan_date_index, value=datetime.now())

        wb.save(template_path)
        print(f"Data added successfully for account number {account_number}.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
account_number = input("Enter the account number: ")
form_data = {
    'Loan Application Date': '2024-03-03',
    'Purpose of the Loan': 'for Agriculture',
    'Address/Location': 'Masindi',
    'Business Financed': 'Farming',
    'Group Name': 'Masindi Youth Farmers',
    'Reason for Default (Summarised)': 'Financial hardship',
    'Detailed Reason for Default': 'Animals died'
}

populate_excel(account_number, form_data)

